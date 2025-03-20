from typing import Union
from fastapi import FastAPI, File, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
import pandas as pd
import uuid
import re
import os
from openpyxl import load_workbook
import sort_in_excel
# from decodeFile import getTextFormat
app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# @app.get("/add")
# def read_root():
#     return {"Hello": "World"}


@app.post("/uploadfile/")
def transcrib_sound_file(file: UploadFile = File(...)):
    try:

        mockData = '[0:00:00.000 --> 0:00:02.080] Стоматология, все свои, здравствуйте. [0:00:04.280 --> 0:00:07.640] Здравствуйте, а соедините меня с Войковской. [0:00:08.439 --> 0:00:10.300] Когда вы были в нашей клинике в последний раз? [0:00:10.560 --> 0:00:11.339] Две недели назад. [0:00:12.300 --> 0:00:15.339] Составайтесь на линии, пожалуйста, вас соединю с клиникой напрямую.'

        # return {"filename": getTextFormat(file.file), "status": "success"}
        return {"filename": mockData, "status": "success"}
    except Exception as e:
        return {"error": str(e)}
    

@app.post("/othercalls/")
async def  analyze_excel_file(file: UploadFile = File(..., pattern=r".*\.xlsx$"),
                              output_file: str = Form("result")): 
    try:
        safe_file = re.sub(r'[\\/:*?"<>|]', '_', output_file)
        safe_file = safe_file[:50].strip()
        
        if not safe_file:
            safe_file = "result"
            
        if not safe_file.endswith('.xlsx'):
            safe_file += '.xlsx'

        # Обработка файла
        wb = load_workbook(file)
        ws = wb.active
        
        # Собираем URL из гиперссылок начиная с 4-й строки
        urls = []
        
        for row in ws.iter_rows(min_row=4, min_col=7, max_col=7):
            cell = row[0]
            if cell.hyperlink:
                urls.append(cell.hyperlink.target)  
    # Обработка аудио
        results = []
        for url in urls:  #tqdm создает прогресс-бар
            if url and isinstance(url, str):
                results.append(sort_in_excel.process_audio(url))
            else:
                results.append("недействительная ссылка")

        df = pd.read_excel(file, header=2)
        df["status"] = results

        output_path = f"/{safe_file}"
        df.to_excel(output_path, index=False)
        
        return FileResponse(
            output_path,
            filename=safe_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    finally:
        if file and os.path.exists(file):
            os.unlink(file)
        if output_path and os.path.exists(output_path):
            os.unlink(output_path)
        for f in [file, output_file]:
            if os.path.exists(f):
                os.remove(f)
