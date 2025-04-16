
from fastapi import APIRouter, File, UploadFile
from fastapi.responses import FileResponse
import pandas as pd
import re
import os
from openpyxl import load_workbook
from repositories.othercalls import process_audio

router = APIRouter(prefix="/othercalls", tags=["othercalls"])

@router.post("",
             responses={400: {"description": "Bad request"}})
def  analyze_excel_file(file: UploadFile = File(...),
                              output_file: str = None): 
    
    output_path = None
    try:
        if output_file:
            safe_file = re.sub(r'[\\/:*?"<>|]', '_', output_file)
            safe_file = safe_file[:50].strip()
        
        else:
            safe_file = "result"
            
        if not safe_file.endswith('.xlsx'):
            safe_file += '.xlsx'

        # Обработка файла
        wb = load_workbook('./test.xlsx')
        ws = wb.active
        
        # Собираем URL из гиперссылок начиная с 4-й строки
        urls = []
        
        for row in ws.iter_rows(min_row=4, min_col=7, max_col=7):
            cell = row[0]
            if cell.hyperlink:
                urls.append(cell.hyperlink.target)  
    # Обработка аудио
        results = []
        for url in urls:  #tqdm создает прогресс-бар
            if url and isinstance(url, str):
                results.append(process_audio(url))
            else:
                results.append("недействительная ссылка")

        df = pd.read_excel(file, header=2)
        df["status"] = results

        output_path = f"/{safe_file}"
        df.to_excel(output_path, index=False)
        
        return FileResponse(
            output_path,
            filename=safe_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    finally:
        if file and os.path.exists(file):
            os.unlink(file)
        if output_path and os.path.exists(output_path):
            os.unlink(output_path)
        for f in [file, output_file]:
            if os.path.exists(f):
                os.remove(f)
