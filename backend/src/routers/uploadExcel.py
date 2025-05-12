
from fastapi import APIRouter, File, UploadFile, HTTPException, status
from fastapi.responses import FileResponse
import pandas as pd
import re
import tempfile
import os
from werkzeug.http import parse_options_header
from openpyxl import load_workbook
from repositories.uploadExcel import process_audio, EventSourceResponse, status_stream

router = APIRouter(prefix="/othercalls", tags=["othercalls"])

@router.post("",
             responses={400: {"description": "Bad request"}})
async def  analyze_excel_file(file: UploadFile = File(...),
                              output_file: str = None): 

    temp_path = None
    output_path = None
    try:
        if output_file:
            safe_file = re.sub(r'[\\/:*?"<>|]', '_', output_file)
            safe_file = safe_file[:50].strip()
        
        else:
            safe_file = "result"
            
        if not safe_file.endswith('.xlsx'):
            safe_file += '.xlsx'

        file_content = await file.read()
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
            temp_file.write(file_content)
            temp_path = temp_file.name
        # Обработка файла
        wb = load_workbook(temp_path)
        ws = wb.active
        

        ws.cell(row=3, column=13, value="Status") 


        for row_idx, row in enumerate(ws.iter_rows(min_row=4, min_col=12, max_col=12), start=4):
            cell = row[0]
            if cell.hyperlink:
                url = cell.hyperlink.target 
                result = process_audio(url)
            else:
                result = "недействительная ссылка"
            
            # Сохраняем результат в новый столбец
            ws.cell(row=row_idx, column=13, value=result)
        

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as output_path:
            wb.save(output_path.name)
            output_path = output_path.name

        wb.close()
        
        return FileResponse(
            output_path,
            filename=safe_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"  
        )
    
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Ошибка обработки файла: {str(e)}"
        )
    finally:
        # Clean up temporary files
        if temp_path and os.path.exists(temp_path):
            os.remove(temp_path)

@router.get("/stream/{task_id}")
async def stream_status(task_id: str):
    return EventSourceResponse(status_stream(task_id))