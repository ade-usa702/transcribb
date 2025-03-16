# import os
# import uuid
# import requests
# from bs4 import BeautifulSoup
# from urllib.parse import urljoin
# from faster_whisper import WhisperModel
# from pydub import AudioSegment
# import pandas as pd
# from tqdm import tqdm
# from openpyxl import load_workbook

# # Конфигурация
# AUDIO_TEMP_DIR = "temp_audio"
# KEYWORDS = ["в архиве", "архивный", "впервые", "не был", "не была", "не были", "давно",
#             "больше года", "архив", "первичная консультация",
#             "консультация", "не было", "первичное"]
# os.makedirs(AUDIO_TEMP_DIR, exist_ok=True)

# # Загрузка модели
# model = WhisperModel("large", device="cpu", compute_type="int8")

# def check_keywords(text):
#     """Проверяет наличие ключевых слов в тексте"""
#     return any(keyword in text.lower() for keyword in KEYWORDS) if text else False

# def process_audio(url):
#     """Обработка аудио с URL"""
#     try:
#         # Получение страницы с аудио
#         response = requests.get(url, timeout=10)
#         response.raise_for_status()
        
#         # Поиск аудио-ссылок
#         if response.status_code == 200:
#             soup = BeautifulSoup(response.content, 'html.parser')
#             audio_links = []
#             for source in soup.find_all('source', src=True):
#                 if source['src']:
#                     audio_links.append(source['src'])
            
        
#         if not audio_links:
#             return "нет аудио"
        
#         # Скачивание последнего аудиофайла
#         last_audio_url = audio_links[-1]
#         filename = os.path.join(AUDIO_TEMP_DIR, f"{uuid.uuid4()}.wav")
        
#         audio_content = requests.get(last_audio_url, timeout=10).content
#         with open(filename, 'wb') as f:
#             f.write(audio_content)
        
#         # Конвертация аудио
#         audio = AudioSegment.from_file(filename)
#         duration = audio.duration_seconds
        
#         # Транскрибация
#         segments, _ = model.transcribe(filename, language="ru")
#         text = " ".join(segment.text for segment in segments)
        
#         # Определение статуса
#         if duration > 50 and check_keywords(text):
#             return "требует проверки"
#         return "верно"
    
#     except Exception as e:
#         return f"ошибка: {str(e)}"

# def analyze_excel_file(input_path):
#     """Анализ Excel файла"""
#     # Загрузка книги через openpyxl для работы с гиперссылками
#     wb = load_workbook(input_path)
#     ws = wb.active
    
#     # Собираем URL из гиперссылок начиная с 4-й строки
#     urls = []
#     for row in ws.iter_rows(min_row=4, min_col=7, max_col=7):
#         cell = row[0]
#         if cell.hyperlink:
#             urls.append(cell.hyperlink.target)  
#             #print(cell.hyperlink.target)# На случай если ссылка в тексте
    
#     # Обработка аудио
#     results = []
#     for url in tqdm(urls, desc="Обработка"):  #tqdm создает прогресс-бар
#         if url and isinstance(url, str):
#             results.append(process_audio(url))
#             print(process_audio(url))
#         else:
#             results.append("недействительная ссылка")
    
#     # Создаем DataFrame и сохраняем результат
#     df = pd.read_excel(input_path, header=2)
#     df['Results'] = results
    
#     output_file = "processed_result.xlsx"
#     df.to_excel(output_file, index=False)
#     print(f"Результаты сохранены в {output_file}")

# # Пример использования
# if __name__ == "__main__":
#     # Создайте тестовый Excel-файл с колонкой "Запись разговора" и ссылками
#     analyze_excel_file("test.xlsx")


import io
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from faster_whisper import WhisperModel
from pydub import AudioSegment
import pandas as pd
from tqdm import tqdm
from openpyxl import load_workbook

# Конфигурация
KEYWORDS = ["в архиве", "архивный", "впервые", "не был", "не была", "не были", "давно",
            "больше года", "архив", "первичная консультация", "консультация", "не было", "первичное"]

# Загрузка модели
model = WhisperModel("large", device="cpu", compute_type="int8")

def check_keywords(text):
    """Проверяет наличие ключевых слов в тексте"""
    return any(keyword in text.lower() for keyword in KEYWORDS) if text else False

def w_keywords(text):
    if not text:  # Проверка на пустой текст
        return []
    
    text_lower = text.lower()
    return [keyword for keyword in KEYWORDS if f" {keyword} " in f" {text_lower} "]

def process_audio(url):
    """Обработка аудио с URL"""
    try:
        # Получение страницы с аудио
        response = requests.get(url, timeout=20)
        response.raise_for_status()
        
        # Поиск аудио-ссылок
        soup = BeautifulSoup(response.content, 'html.parser')
        audio_links = []
        for source in soup.find_all('source', src=True):
            if source['src']:
                audio_links.append(source['src'])
        
        if not audio_links:
            return "нет аудио"
        
        # Скачивание аудио в память
        audio_response = requests.get(audio_links[-1], timeout=20)
        audio_response.raise_for_status()
        
        # Обработка аудио в памяти
        with io.BytesIO(audio_response.content) as audio_buffer:
            # Конвертация в формат WAV
            audio = AudioSegment.from_file(audio_buffer)
            audio = audio.set_frame_rate(16000).set_channels(1)
            
            # Конвертация в сырые данные для Whisper
            with io.BytesIO() as wav_buffer:
                audio.export(wav_buffer, format="wav")
                wav_buffer.seek(0)
                
                # Транскрибация
                segments, _ = model.transcribe(wav_buffer, language="ru")
                text = " ".join(segment.text for segment in segments)
        
        # Определение статуса
        duration = len(audio) / 1000  # Продолжительность в секундах
        if duration < 50:
             status = "длительность менее 50 секунд"
        elif duration > 50 and check_keywords(text):
            status = "требует проверки"
        else:
            status = "верно"

        details = f" ({w_keywords(text)})" if status == "верно" else ""
        return f"{status}{details}"
        
    except Exception as e:
        return f"ошибка: {str(e)}"

def analyze_excel_file(input_path):
    """Анализ Excel файла"""
    # Загрузка книги через openpyxl для работы с гиперссылками
    wb = load_workbook(input_path)
    ws = wb.active
    
    # Собираем URL из гиперссылок начиная с 4-й строки
    urls = []
    for row in ws.iter_rows(min_row=4, min_col=7, max_col=7):
        cell = row[0]
        if cell.hyperlink:
            urls.append(cell.hyperlink.target)  
            #print(cell.hyperlink.target)# На случай если ссылка в тексте
    
    # Обработка аудио
    results = []
    for url in tqdm(urls, desc="Обработка"):  #tqdm создает прогресс-бар
        if url and isinstance(url, str):
            results.append(process_audio(url))
            print(process_audio(url))
        else:
            results.append("недействительная ссылка")
    
    # Создаем DataFrame и сохраняем результат
    df = pd.read_excel(input_path, header=2)
    df['Results'] = results
    
    output_file = "processed_result.xlsx"
    df.to_excel(output_file, index=False)
    print(f"Результаты сохранены в {output_file}")


if __name__ == "__main__": 
    analyze_excel_file("test.xlsx")
    